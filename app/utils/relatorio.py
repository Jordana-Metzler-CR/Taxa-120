from dataclasses import dataclass, field
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


@dataclass
class LinhaRelatorio:
    cod_imovel: str
    numero_taxa: str
    descricao_taxa: str
    valor: str
    processo: str
    status: str
    mensagem: str


@dataclass
class RelatorioOperacional:
    linhas: List[LinhaRelatorio] = field(default_factory=list)
    _boletos_processados: set = field(default_factory=set, repr=False)
    _boletos_com_erro: set = field(default_factory=set, repr=False)

    def registrar(
        self,
        cod_imovel,
        numero_taxa,
        descricao_taxa: str,
        valor,
        status: str,
        mensagem: str,
        processo: str = "",
    ):
        if valor not in (None, ""):
            try:
                valor_fmt = f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except (ValueError, TypeError):
                valor_fmt = str(valor)
        else:
            valor_fmt = ""

        self.linhas.append(LinhaRelatorio(
            cod_imovel=str(cod_imovel) if cod_imovel else "",
            numero_taxa=str(numero_taxa) if numero_taxa else "",
            descricao_taxa=str(descricao_taxa) if descricao_taxa else "",
            valor=valor_fmt,
            processo=processo,
            status=status,
            mensagem=str(mensagem) if mensagem else "",
        ))
        imovel_key = str(cod_imovel) if cod_imovel else ""
        if imovel_key:
            self._boletos_processados.add(imovel_key)
            if status in ("Erro", "Alerta"):
                self._boletos_com_erro.add(imovel_key)

    @property
    def total_boletos(self) -> int:
        return len(self._boletos_processados)

    @property
    def boletos_sucesso(self) -> int:
        return len(self._boletos_processados - self._boletos_com_erro)

    @property
    def boletos_erro(self) -> int:
        return len(self._boletos_com_erro)

    def gerar_excel(self, nome_arquivo: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Taxa 120"

        cabecalhos = [
            "Cód. Imóvel",
            "Nº da Taxa",
            "Descrição da Taxa",
            "Valor (R$)",
            "Processo",
            "Status",
            "Mensagem", 
        ]

        for col_idx, nome in enumerate(cabecalhos, start=1):
            ws.cell(row=1, column=col_idx, value=nome)

        for i, linha in enumerate(self.linhas, start=2):
            ws.cell(row=i, column=1, value=linha.cod_imovel)
            ws.cell(row=i, column=2, value=linha.numero_taxa)
            ws.cell(row=i, column=3, value=linha.descricao_taxa)
            ws.cell(row=i, column=4, value=linha.valor)
            ws.cell(row=i, column=5, value=linha.processo)
            ws.cell(row=i, column=6, value=linha.status)
            ws.cell(row=i, column=7, value=linha.mensagem)

        linha_rodape = len(self.linhas) + 3
        totais = [
            ("Nº de boletos processados",             self.total_boletos),
            ("Nº de boletos processados com sucesso", self.boletos_sucesso),
            ("Nº de boletos com erro",                self.boletos_erro),
        ]
        
        
        for offset, (descricao, valor) in enumerate(totais):
            ws.cell(row=linha_rodape + offset, column=1, value=descricao)
            ws.cell(row=linha_rodape + offset, column=2, value=valor)

        
         # largura dinâmica + alinhamento centralizado em todas as células
        for col_cells in ws.iter_cols():
            largura = max(
                len(str(cel.value)) for cel in col_cells if cel.value is not None
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = largura + 2
 
        for row in ws.iter_rows():
            for cel in row:
                cel.alignment = Alignment(horizontal="center", vertical="center")

        # rodapé
       
        wb.save(nome_arquivo)
        return nome_arquivo